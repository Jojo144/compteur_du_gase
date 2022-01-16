from tempfile import NamedTemporaryFile

from django.db.models import Sum, F
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import *
from .templatetags.my_tags import *


def format_member(m):
    if m.email:
        if m.tel:
            paren = " ({}, {})".format(m.email, m.tel)
        else:
            paren = " ({})".format(m.email)
    else:
        if m.tel:
            paren = " ({})".format(m.tel)
        else:
            paren = ""
    return str(m) + paren


def generate_export_households(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Foyers"
    ws['A1'] = "Foyer"
    ws['B1'] = "Membres"
    ws['C1'] = "Adresse"
    ws['D1'] = "Date d'adhésion"
    ws.row_dimensions[1].font = Font(bold=True)
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20

    for h in Household.objects.all():
        members = ', '.join([format_member(m) for m in Member.objects.filter(household=h)])
        ws.append((h.name, members, h.address, h.date))

    # moche
    for i,_ in enumerate(ws.rows):
        ws.row_dimensions[i+1].height=20

    wb.save(filename)


def generate_export_products(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Produits"
    ws['A1'] = "Nom"
    ws['B1'] = "Catégorie"
    ws['C1'] = "Fournisseur"
    ws['D1'] = "Prix"
    ws['E1'] = "Visible"
    ws['F1'] = "Unité"
    ws['G1'] = "Prix libre"
    ws['H1'] = "Stock"
    ws.row_dimensions[1].font = Font(bold=True)
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['F'].width = 20

    for p in Product.objects.all():
        ws.append((p.name, str(p.category), str(p.provider), p.price, bool_to_utf8(p.visible), str(p.unit), bool_to_utf8(p.pwyw), p.stock))

    # moche
    for i,_ in enumerate(ws.rows):
        ws.row_dimensions[i+1].height=20

    wb.save(filename)


def generate_export_providers(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fournisseurs"
    ws['A1'] = "Nom"
    ws['B1'] = "Contact"
    ws['C1'] = "Commentaire"
    ws.row_dimensions[1].font = Font(bold=True)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 60

    for p in Provider.objects.all():
        ws.append((p.name, p.contact, p.comment))

    # moche
    for i,_ in enumerate(ws.rows):
        ws.row_dimensions[i+1].height=35

    wb.save(filename)


def generate_ecarts_data():
    ope = ChangeStockOp.objects.filter(label='Inventaire')
    stats = ope.annotate(
        calendar_date=F('date__date')
    ).values(
        'calendar_date'
    ).annotate(
        total=Sum('price')
    )[:10]  # 10 dernières dates

    ecarts_data = [{'date': stat['calendar_date'],
                    'details': ope.filter(date__date=stat['calendar_date']),
                    'result': stat['total']}
                   for stat in stats]
    return ecarts_data

def generate_export_inventory(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan des 10 derniers inventaires"
    ws['A1'] = "Date"
    ws['B1'] = "Produit"
    ws['C1'] = "Quantité"
    ws['D1'] = "Valeur (en €)"
    ws.row_dimensions[1].font = Font(bold=True)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    for d in generate_ecarts_data():
        ws.append((d['date'],))
        for op in d['details']:
            ws.append((None, str(op.product), print_quantity_op(op), op.price))
        ws.append((None, 'Total', None, d['result']))

    # moche
    for i,_ in enumerate(ws.rows):
        ws.row_dimensions[i+1].height=20

    wb.save(filename)
